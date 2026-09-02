# =============================================================================
# htr_sync.py  --  importable module, writes nothing
# -----------------------------------------------------------------------------
# PURPOSE
#   Provides the video-frame -> photometry-time mapping as a FUNCTION, so the
#   analysis scripts get corrected onsets in memory. No intermediate CSV, which
#   sidesteps the OneDrive/Defender write problem entirely.
#
# USAGE 
# Maps manually scored video-frame indices to photometry timestamps using the 
# recorded camera synchronization TTL train.
# 
# Direct TTL-based mapping accounts for small deviations from the nominal camera
# frame rate that would otherwise accumulate as timing drift over long recordings
#
# =============================================================================

import numpy as np
import pandas as pd

# --- photometry file layout ---
COL_TIME = 0     # A: Time(s)
SYNC_COL = 4     # E: DI/O #2  -- camera frame TTL (20 Hz)
                 # NOTE col 1 (CAM #1) is the 40 Hz photometry exposure strobe,
                 # NOT the behavioural camera. Never use it here.

# --- scoring sheet layout (0-indexed) ---
# SCORE_NOTE_COL assumes the standard sheet layout (notes in column H). If a
# sheet ever arrives with a different layout, the note inventory printed by
# load_scoring() is what tells you -- check it every run.
SCORE_INDEX_COL, SCORE_FRAME_COL, SCORE_NOTE_COL = 2, 5, 7

NOMINAL_FPS = 20.0    # what the AVI container claims / what the sheet assumed

_CACHE = {}           # in-memory only; avoids re-reading 5M rows per script run


def frame_times(photometry_path, sync_col=SYNC_COL):
    """Photometry timestamp of every camera frame: frame_times[N] = time of
    video frame N. Reads only the two needed columns (the file is ~5M rows).
    Cached in memory for the life of the process."""
    key = (photometry_path, sync_col)
    if key in _CACHE:
        return _CACHE[key]

    raw = pd.read_csv(photometry_path, header=None, low_memory=False,
                      usecols=sorted({COL_TIME, sync_col}), skiprows=2)
    raw = raw.apply(pd.to_numeric, errors="coerce")
    t = raw[COL_TIME].to_numpy()
    # digital lines hold state across blank rows -> forward-fill before diffing
    x = raw[sync_col].ffill().fillna(0).to_numpy()
    idx = np.where((x[1:] == 1) & (x[:-1] == 0))[0] + 1
    ft = t[idx]

    _CACHE[key] = ft
    return ft


def report_sync(photometry_path, sync_col=SYNC_COL, video_n_frames=None):
    """Re-run the validation. Call once per session; prints, returns a dict.
    RAISES if the sync train has gaps -- see the comment at the check."""
    ft = frame_times(photometry_path, sync_col)
    d = np.diff(ft)
    med = np.median(d)
    mean_int = (ft[-1] - ft[0]) / (len(ft) - 1)
    n_gaps = int(np.sum(d > 1.5 * med))

    print(f"sync col {sync_col}: {len(ft):,} edges | {ft[0]:.4f} - {ft[-1]:.4f} s")
    print(f"  mean interval {mean_int*1000:.6f} ms -> TRUE {1/mean_int:.5f} Hz "
          f"(nominal {NOMINAL_FPS})")
    print(f"  gaps (>1.5x median): {n_gaps}"
          + ("  [ok] index N == frame N is safe" if n_gaps == 0
             else "  [!] INDEXING UNSAFE after the first gap"))
    if n_gaps:
        raise ValueError(f"{n_gaps} gap(s) in the sync train -- frame index != edge "
                         f"index after the first gap. Onsets are not trustworthy.")
    if video_n_frames is not None:
        print(f"  video frames {video_n_frames:,} - edges {len(ft):,} = "
              f"{video_n_frames - len(ft):+,} "
              f"({abs(video_n_frames - len(ft))/NOMINAL_FPS:.2f} s unaccounted)")
    return {"n_edges": len(ft), "mean_interval": mean_int,
            "true_fps": 1/mean_int, "n_gaps": n_gaps}


def load_scoring(scoring_path):
    """Accept a row only if BOTH the HTR index and the frame are numeric, so a
    stray numeric cell elsewhere in the sheet cannot become a phantom twitch.
    Prints anything rejected rather than dropping it silently.

    Also prints the distinct scorer notes found. That printout is the safeguard
    against a wrong SCORE_NOTE_COL: pointing at an empty column raises nothing,
    it just makes EXCLUDE_FLAGGED do nothing and lets questionable twitches into
    the average. Every other failure here announces itself with a traceback."""
    from openpyxl import load_workbook
    ws = load_workbook(scoring_path, read_only=True, data_only=False).active
    keep, rejected = [], []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        idx   = r[SCORE_INDEX_COL] if len(r) > SCORE_INDEX_COL else None
        frame = r[SCORE_FRAME_COL] if len(r) > SCORE_FRAME_COL else None
        note  = r[SCORE_NOTE_COL]  if len(r) > SCORE_NOTE_COL  else None
        if isinstance(frame, (int, float)):
            if isinstance(idx, (int, float)):
                keep.append({"htr_idx": int(idx), "frame": int(frame),
                             "note": (str(note).strip() if note not in (None, "") else "")})
            else:
                rejected.append((i, frame))
    if rejected:
        print("[!] numeric frame but no HTR index -- NOT treated as twitches:")
        for i, f in rejected:
            print(f"    sheet row {i}: frame {f}")

    notes = sorted({k["note"] for k in keep if k["note"]})
    n_noted = sum(1 for k in keep if k["note"])
    print(f"scoring: {len(keep)} rows | {n_noted} carry a note")
    for n in notes:
        print(f"    note: {n!r}")
    if not notes:
        print("    [!] NO notes found -- either the sheet has none, or "
              "SCORE_NOTE_COL is wrong and EXCLUDE_FLAGGED is doing nothing.")
    if any(n.startswith("=") for n in notes):
        raise ValueError(f"SCORE_NOTE_COL={SCORE_NOTE_COL} holds formulas, not "
                         f"notes. Check the sheet layout before continuing.")
    return pd.DataFrame(keep)


def corrected_onsets(photometry_path, scoring_path, frame_offset=0,
                     sync_col=SYNC_COL, verbose=True):
    """THE function to call. Returns a DataFrame sorted by corrected onset:
         htr_idx, frame, onset_naive_s, onset_s, shift_s, note
       onset_s is the corrected photometry time. Unmappable frames are dropped
       (and reported)."""
    ft = frame_times(photometry_path, sync_col)
    ev = load_scoring(scoring_path)
    if ev.empty:
        raise ValueError("No scored HTRs found -- check the scoring column indices.")

    k = ev["frame"].values - frame_offset
    ok = (k >= 0) & (k < len(ft))
    if verbose and (~ok).any():
        for _, r in ev[~ok].iterrows():
            print(f"[!] HTR {r['htr_idx']} frame {r['frame']} outside sync train "
                  f"(0..{len(ft)-1}) -- dropped")

    ev = ev[ok].copy()
    ev["onset_naive_s"] = ev["frame"] / NOMINAL_FPS
    ev["onset_s"]       = ft[ev["frame"].values - frame_offset]
    ev["shift_s"]       = ev["onset_s"] - ev["onset_naive_s"]
    ev = ev.sort_values("onset_s").reset_index(drop=True)

    if verbose:
        print(f"mapped {len(ev)} twitches | frames {ev['frame'].min():,}-"
              f"{ev['frame'].max():,} | shift {ev['shift_s'].min():+.3f} to "
              f"{ev['shift_s'].max():+.3f} s (drift "
              f"{ev['shift_s'].max()-ev['shift_s'].min():+.3f} s)")
    return ev


if __name__ == "__main__":
    # standalone check -- edit these two paths to whichever session you want
    PHOTO = r"example_photometry.csv"
    SCORE = r"example.scoring.xlsx"
    report_sync(PHOTO)
    ev = corrected_onsets(PHOTO, SCORE)
    print(ev.head(10).to_string(index=False))
    
    